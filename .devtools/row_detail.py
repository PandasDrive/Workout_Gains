from openpyxl import load_workbook
wb = load_workbook('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx', data_only=True)
ws = wb['Intermediate-Advanced Program']
row = 30
values = []
for col in range(1, 21):
    cell = ws.cell(row=row, column=col)
    val = cell.value
    values.append((col, val))
print('Row', row)
for col, val in values:
    print(col, val)
