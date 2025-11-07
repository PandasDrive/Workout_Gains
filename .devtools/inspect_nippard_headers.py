from openpyxl import load_workbook
import pathlib
wb = load_workbook('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx', data_only=True)
ws = wb['Intermediate-Advanced Program']
for idx, row in enumerate(ws.iter_rows(min_row=28, max_row=34, values_only=True), start=28):
    items = [f"{chr(65+i)}:{'' if cell is None else cell}" for i, cell in enumerate(row[:15])]
    print(idx, ' | ', ' | '.join(items))
