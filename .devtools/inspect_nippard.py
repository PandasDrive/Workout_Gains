from openpyxl import load_workbook
import pathlib
wb_path = pathlib.Path('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx')
wb = load_workbook(wb_path, data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
    line = '\t'.join('' if cell is None else str(cell) for cell in row)
    print(line.encode('ascii', errors='ignore').decode())
