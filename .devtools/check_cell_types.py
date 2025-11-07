from openpyxl import load_workbook
wb = load_workbook('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx', data_only=True)
ws = wb['Intermediate-Advanced Program']
cell = ws['E30']
print('E30 value:', cell.value, type(cell.value))
cell = ws['G30']
print('G30 value:', cell.value, type(cell.value))
cell = ws['E31']
print('E31 value:', cell.value, type(cell.value))
