from openpyxl import load_workbook
from datetime import datetime
import json

wb = load_workbook('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx', data_only=True)
ws = wb['Intermediate-Advanced Program']

week_markers = []
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=2).value
    if isinstance(val, str) and val.strip().lower().startswith('week '):
        week_markers.append((row, val.strip()))

week_markers.append((ws.max_row + 1, None))

def format_range(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text

def clean_text(value):
    if value is None:
        return ""
    text = str(value)
    replacements = {
        '\n': ' ',
        '°': ' deg',
        '–': '-',
        '’': "'",
        '“': '"',
        '”': '"'
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = text.strip()
    while '  ' in text:
        text = text.replace('  ', ' ')
    return text

weeks_data = []
for idx in range(len(week_markers) - 1):
    start_row, week_label = week_markers[idx]
    end_row = week_markers[idx + 1][0]
    week_number = int(week_label.split()[1])
    days = []
    current_day = None
    for row in range(start_row + 2, end_row):
        day_label = clean_text(ws.cell(row=row, column=2).value)
        exercise_name = clean_text(ws.cell(row=row, column=3).value)
        rest_marker = clean_text(ws.cell(row=row, column=10).value)

        if day_label:
            current_day = {
                'day': len(days) + 1,
                'title': day_label,
                'cardio': '',
                'exercises': []
            }
            days.append(current_day)
        if rest_marker and rest_marker.lower() == 'rest day':
            rest_day = {
                'day': len(days) + 1,
                'title': 'Rest Day',
                'cardio': '',
                'exercises': []
            }
            days.append(rest_day)
            current_day = None
            continue
        if exercise_name and current_day:
            warmup = format_range(ws.cell(row=row, column=5).value)
            working_sets = format_range(ws.cell(row=row, column=6).value)
            reps = format_range(ws.cell(row=row, column=7).value)
            early_rpe = clean_text(ws.cell(row=row, column=12).value)
            last_rpe = clean_text(ws.cell(row=row, column=13).value)
            rest_text = clean_text(ws.cell(row=row, column=14).value)
            sub1 = clean_text(ws.cell(row=row, column=15).value)
            sub2 = clean_text(ws.cell(row=row, column=16).value)
            note_extra = clean_text(ws.cell(row=row, column=17).value)

            notes_parts = []
            if warmup and warmup.lower() != 'n/a':
                notes_parts.append(f"Warm-up sets: {warmup}")
            if early_rpe and early_rpe.lower() != 'n/a':
                notes_parts.append(f"Early set RPE: {early_rpe}")
            if last_rpe and last_rpe.lower() != 'n/a':
                notes_parts.append(f"Last set RPE: {last_rpe}")
            if sub1:
                notes_parts.append(f"Sub option 1: {sub1}")
            if sub2:
                notes_parts.append(f"Sub option 2: {sub2}")
            if note_extra:
                notes_parts.append(note_extra)
            notes = ' | '.join(notes_parts)

            current_day['exercises'].append({
                'name': exercise_name,
                'sets': working_sets,
                'reps': reps,
                'rest': rest_text,
                'notes': notes
            })
    weeks_data.append({
        'week': week_number,
        'task': '',
        'days': days
    })

with open('.devtools/nippard_data.json', 'w', encoding='utf-8') as f:
    json.dump(weeks_data, f, ensure_ascii=False, indent=2)

print('Generated data for', len(weeks_data), 'weeks')
