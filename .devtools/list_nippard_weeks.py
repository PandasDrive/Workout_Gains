from openpyxl import load_workbook
wb = load_workbook('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx', data_only=True)
ws = wb['Intermediate-Advanced Program']
weeks = []
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=2).value
    if isinstance(val, str) and val.strip().lower().startswith('week '):
        weeks.append((row, val.strip()))
print('Weeks found:', len(weeks))
for info in weeks:
    print(info)
