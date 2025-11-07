from openpyxl import load_workbook
import pathlib
wb_path = pathlib.Path('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx')
wb = load_workbook(wb_path, data_only=True)
ws = wb['Intermediate-Advanced Program']
for idx, row in enumerate(ws.iter_rows(min_row=35, max_row=60, values_only=True), start=35):
    items = [f"{chr(65+i)}:{'' if cell is None else cell}" for i, cell in enumerate(row[:12])]
    print(idx, ' | ', ' | '.join(items))
