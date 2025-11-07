from openpyxl import load_workbook
wb = load_workbook('WorkoutPrograms/The_Bodybuilding_Transformation_System_-_Intermediate-Advanced.xlsx', data_only=True)
ws = wb['Intermediate-Advanced Program']
row = 43
vals = []
for col in range(1, 21):
    vals.append((col, ws.cell(row=row, column=col).value))
print(vals)
